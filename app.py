#!/usr/bin/env python3
"""
INFINITE GZ + AI SMART TECH — Invoice Management System  v3 (patch-v12.5)
=============================================================
Changes vs v2:
  • IGZ: Items (JSON) → 4 cols: Product Item / Qty / Unit Price (RM) / Total Amount (RM)
  • IGZ Product Item dropdown: 51 transaction types from IGZ Transaction Reference doc
  • IGZ 2026 market reference prices added alongside each transaction type
  • Both companies: Payment Type & Card Type dropdowns in Excel + UI
  • Screen 1 right column: shows BOTH company price reference tables (IGZ top, AST bottom)
  • Receipt Link & Invoice Link: wide columns + horizontal scroll
  • Invoice No: uses actual receipt invoice number (INV- prefix stripped automatically)
"""

import streamlit as st
import pytesseract
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import re, os, io, base64, datetime, subprocess, json
from pathlib import Path

import sys
sys.path.insert(0, str(Path(__file__).parent))
from invoice_infinitegz_v7 import make_igz
from invoice_ast import make_ast

# ── Paths (with /mnt fallback to local) ────────────────────────────────
def _safe_base() -> Path:
    """Use ./output/ relative to script location — works on any OS."""
    base = Path(__file__).parent / "output"
    base.mkdir(parents=True, exist_ok=True)
    return base

OUT_DIR       = _safe_base()
IGZ_EXCEL     = OUT_DIR / "IGZ_invoices.xlsx"
AST_EXCEL     = OUT_DIR / "AST_invoices.xlsx"
RECEIPTS_DIR  = OUT_DIR / "receipts"
INVOICES_DIR  = OUT_DIR / "invoices"
for d in [RECEIPTS_DIR, INVOICES_DIR]:
    d.mkdir(parents=True, exist_ok=True)

CARD_MAP_FILE = OUT_DIR / "card_map.json"   # card last-4 → Bill To name

# ── CC Management seed (from CC MANAGEMENT.xlsx – 61 entries) ────────────────
# Automatically pre-seeds the card_map so Bill To is resolved without
# manual setup on a fresh Streamlit Cloud deployment.
CC_SEED: dict = {
      "0066": "TAN YONG SHENG",
      "0140": "CARRIE TONG",
      "0235": "ANGELINE LEE POI LIN",
      "0446": "CHIA VUI LEONG",
      "0555": "SONG YEW CHUAN",
      "0670": "HASAN",
      "1259": "HASRUL",
      "1470": "LAI FOOK HENG",
      "1513": "CHIA VUI LEONG",
      "1560": "CHIA VUI LEONG",
      "1818": "CHOW KAH FEI",
      "2033": "CHANG CHOON CHOW",
      "2058": "CHANG CHOON CHOW",
      "2322": "CHIA VUI LEONG",
      "2385": "LEE E KAI",
      "2404": "LYE PEI KUN",
      "2530": "HASAN",
      "2682": "YOONG MENG WEE",
      "2978": "CHIA VUI LEONG",
      "3041": "CARRIE TONG",
      "3123": "GOH MUI HIM",
      "3447": "OOI CHIEW FOONG",
      "3576": "CHIA VUI LEONG",
      "3687": "KHOR WEI LOONG",
      "3717": "CHIA VUI LEONG",
      "3770": "CHOW KAH FEI",
      "3964": "LYE PEI KUN",
      "3998": "LEE E KAI",
      "4127": "LEOW HOOI SZE",
      "4511": "CHIA VUI LEONG",
      "4514": "CHANG CHOON CHOW",
      "4523": "YEO CHUAN TZUN",
      "4543": "CHIA VUI LEONG",
      "5474": "SONG YEW CHUAN",
      "5700": "CHIA VUI LEONG",
      "5781": "YEO CHUAN TZUN",
      "6003": "LAI FOOK HENG",
      "6506": "LIM SHAN WEN",
      "6821": "CHOW KAH FEI",
      "6854": "CHIA VUI LEONG",
      "6940": "LIM SHAN WEN",
      "7427": "YEO CHUAN TZUN",
      "7496": "YOONG MENG WEE",
      "7531": "WOO WEN BIN",
      "7698": "CHIA VUI LEONG",
      "7770": "YEO CHUAN TZUN",
      "7809": "TAN YONG SHENG",
      "8036": "GOH MUI HIM",
      "8074": "LEE CHEE HWA",
      "8108": "CHIA VUI KHENG",
      "8114": "CHIA VUI LEONG",
      "8363": "CHIA VUI LEONG",
      "8461": "CHANG CHOON CHOW",
      "8564": "YOONG MENG WEE",
      "8822": "CHEONG HSIU YEING",
      "8887": "LIM SHAN WEN",
      "8894": "CHIA VUI KHENG",
      "9280": "LEOW HOOI SZE",
      "9383": "WOO WEN BIN",
      "9558": "LEE E KAI",
      "9791": "LEE E KAI"
    }


# ══════════════════════════════════════════════════════════════════════
#  INFINITE GZ Product Catalogue  (from IGZ Transaction Reference v1.0)
#  Source: IGZ_Transaction_Reference.docx  |  2026 Malaysia market prices
# ══════════════════════════════════════════════════════════════════════
IGZ_PRODUCTS = [
    # ── GROUP A: Financing Arrangement ──────────────────────────────
    ("[A-001] Card Balance Alignment Package",           1500),
    ("[A-002] Debt Position Adjustment Package",         2000),
    ("[A-003] Credit Profile Care Plan",                  800),
    ("[A-004] Personal Financing Arrangement Pack",      1200),
    ("[A-005] Home Financing Setup Pack",                3000),
    ("[A-006] Business Funding Setup Pack",              5000),
    ("[A-007] Emergency Liquidity Arrangement Pack",     1500),
    ("[A-008] Credit Limit Restructuring Pack",          1200),
    # ── GROUP B: Record & Report Handling ────────────────────────────
    ("[B-001] Monthly Entry Handling Pack",               500),
    ("[B-002] Year-End Report Preparation Pack",         2500),
    ("[B-003] Tax Working File Pack",                    1500),
    ("[B-004] Business Figures Summary Pack",             800),
    ("[B-005] Bank Movement Matching Package",            500),
    ("[B-006] Payroll Processing Pack",                   450),
    ("[B-007] Account Reconciliation Pack",               800),
    ("[B-008] Compliance Documentation Pack",            1500),
    # ── GROUP C: Digital & Business Services ─────────────────────────
    ("[C-001] Website Build & Launch Pack",              5000),
    ("[C-002] Online Reach Boost Package",               3000),
    ("[C-003] Space Layout & Sourcing Pack",             3500),
    ("[C-004] Business Direction Planning Pack",         2500),
    ("[C-005] Document Draft & Filing Pack",              800),
    ("[C-006] Financial Skills Session",                 1200),
    ("[C-007] Brand Identity Setup Pack",                4500),
    ("[C-008] Digital Tools Integration Pack",           3000),
    # ── GROUP D: Outcome Sharing ─────────────────────────────────────
    ("[D-001] Business Setup Integration Pack",          3000),
    ("[D-002] Financing Cost Reduction Outcome",         1000),
    ("[D-003] Extra Income Channel Setup Pack",          2000),
    ("[D-004] Debt Clearing Flow Pack",                  1000),
    ("[D-005] VIP Rotation Access Pack",                 1500),
    ("[D-006] Program Net Outcome Share",                 800),
    ("[D-007] Network Outcome Share",                     800),
    ("[D-008] Product Option Matching Pack",              800),
    ("[D-009] Ongoing Collaboration Outcome Share",       800),
    ("[D-010] Referral Outcome Share",                    500),
    ("[D-011] Agent Payout Share",                        500),
    ("[D-012] POS Terminal MDR Settlement",               200),
    # ── GROUP E: Bank Account Operations ─────────────────────────────
    ("[E-001] Inward Fund Transfer – Bank In",             50),
    ("[E-002] Outward Fund Transfer – Bank Out",           50),
    ("[E-003] IBG Inward – IBG In",                        30),
    ("[E-004] IBG Outward – IBG Out",                      30),
    ("[E-005] Cash Deposit – CASH_DEP",                    50),
    ("[E-006] Cash Withdrawal – CASH_WD",                  50),
    ("[E-007] Cheque Payment – CHEQUE",                    80),
    ("[E-008] Bill Payment – Bill Pay",                    30),
    ("[E-009] Loan Disbursement Received – Loan Disb",    100),
    ("[E-010] Loan Repayment Made – Loan Repay",           80),
    ("[E-011] POS Payment Received – POS Pay",             50),
    ("[E-012] Intercompany Transfer",                     100),
    ("[E-013] MDR Charge Deduction",                       50),
    ("[E-014] Standing Instruction Payment",               50),
    # ── GROUP U: Uncertain / Pending ─────────────────────────────────
    ("[U-001] Pending Review – 待确认",                     0),
]

# ══════════════════════════════════════════════════════════════════════
#  AI SMART TECH Product Catalogue  (2026 Malaysia market prices)
# ══════════════════════════════════════════════════════════════════════
AST_PRODUCTS = [
    ("Website Development – Basic (up to 10 pages)",            3800),
    ("Website Development – Business (custom design)",          8500),
    ("Website Development – E-Commerce",                       18000),
    ("Website Development – Enterprise / Portal",              45000),
    ("Mobile App Development – iOS / Android (basic)",         15000),
    ("Mobile App Development – Custom Full-Stack",             50000),
    ("Software Development – Custom System (SME)",             20000),
    ("Software Development – Enterprise Platform",             80000),
    ("CRM System – Setup & Configuration",                      8000),
    ("CRM System – Custom Development & Integration",          25000),
    ("CRM System – Monthly Support & Maintenance",              1500),
    ("AI Automation – Workflow Bot (single process)",           5000),
    ("AI Automation – Multi-Process Automation Suite",         20000),
    ("AI Automation – Enterprise AI Agent System",             75000),
    ("AI Chatbot – Basic FAQ Bot",                              3500),
    ("AI Chatbot – Advanced NLP Customer Service Bot",         12000),
    ("Hardware Setup – Workstation Configuration (per unit)",    350),
    ("Hardware Setup – Network & Server Infrastructure",        8500),
    ("Hardware Setup – Enterprise IT Infrastructure",          35000),
    ("Brand Strategy & Positioning Package",                    6500),
    ("Brand Identity Design (logo, colours, guidelines)",       4500),
    ("Digital Marketing – Monthly Management (2 platforms)",    4500),
    ("Digital Marketing – Full Campaign Management",            8000),
    ("SEO Optimisation – Monthly Package",                      2500),
    ("Google / Meta Ads Management – Monthly",                  2000),
    ("Social Media Content Creation – Monthly (10 posts)",      2200),
    ("Technical Support – Monthly Retainer (8×5)",              2500),
    ("Technical Support – Monthly Retainer (24×7)",             5000),
    ("Technical Support – Per-Incident On-Site",                 500),
    ("IT Consultation – Per Session (2 hours)",                  800),
    ("Cybersecurity Audit & Assessment",                        6000),
    ("Cloud Migration & Deployment",                           12000),
    ("Data Analytics Dashboard Setup",                          9000),
    ("Microsoft 365 / Google Workspace Deployment",             3500),
    ("Staff IT Training Programme (per session)",               1200),
    ("VPN & Network Security Configuration",                    4500),
    ("2-Year Enterprise Technical Support Package",            18000),
]

# ══════════════════════════════════════════════════════════════════════
#  Package Definitions  (Combination 2: Predefined + Auto Sub-items)
# ══════════════════════════════════════════════════════════════════════

# Format: { "Package Name": ["sub-item 1", "sub-item 2", ...] }

AST_PACKAGES = {
    "": [],
    "Starter Digital Package": [
        "Website Development (Basic)",
        "Brand Identity Design (Logo + Business Card)",
        "Social Media Setup & Optimization",
        "Technical Support – Basic (3 Months)",
    ],
    "Business Growth Package": [
        "Website Development (Corporate)",
        "CRM System – Basic",
        "AI Chatbot Integration – Basic",
        "Digital Marketing Strategy & Consultation",
        "SEO Optimization (On-Page)",
        "Technical Support – Standard (6 Months)",
    ],
    "Enterprise Digital Package": [
        "Website Development (E-Commerce)",
        "Mobile App Development – Basic (iOS + Android)",
        "CRM System – Advanced",
        "AI Chatbot Integration – Advanced",
        "Cloud Infrastructure Setup & Migration",
        "Cybersecurity Audit & Implementation",
        "Technical Support – Premium (12 Months)",
    ],
    "Premium AI Suite": [
        "Custom AI Solution Development",
        "Mobile App Development – Advanced (iOS + Android)",
        "Enterprise CRM System (Full Suite)",
        "AI Automation Workflow Integration",
        "Data Analytics Dashboard",
        "VPN & Network Security Configuration",
        "Dedicated 24/7 Technical Support (24 Months)",
        "Staff Training & Onboarding Program",
    ],
    "Custom Package": [],  # User defines sub-items manually
}

IGZ_PACKAGES = {
    "": [],
    "Financing Arrangement Package": [
        "[A-001] Card Balance Alignment Package",
        "[A-002] Credit Limit Optimisation Service",
        "[A-003] Tax Working File Pack",
        "[A-004] Payment Schedule Management",
        "[A-005] Financial Record Documentation",
    ],
    "Record & Report Management Package": [
        "[B-001] Monthly Statement Processing",
        "[B-002] Transaction Record Management",
        "[B-003] Document Verification Service",
        "[B-004] Compliance Report Generation",
        "[B-005] Financial Summary Preparation",
    ],
    "Digital Business Services Package": [
        "[C-001] Business Registration Assistance",
        "[C-002] Digital Business Profile Setup",
        "[C-003] Document Preparation & Notarisation",
        "[C-004] Compliance Advisory Service",
        "[C-005] Corporate Secretarial Services",
    ],
    "Outcome Sharing & Settlement Package": [
        "[D-001] Revenue Distribution Management",
        "[D-002] Profit Sharing Documentation",
        "[D-003] Performance Review & Reporting",
        "[D-004] Settlement Processing Service",
        "[D-005] Outcome Verification & Sign-off",
    ],
    "Bank Account Operations Package": [
        "[E-001] Bank Account Activation Service",
        "[E-002] Account Maintenance & Monitoring",
        "[E-003] Fund Transfer Facilitation",
        "[E-004] Transaction History Reconciliation",
        "[E-005] Account Health Score Reporting",
    ],
    "Full Service Package": [
        "[A-001] Card Balance Alignment Package",
        "[B-001] Monthly Statement Processing",
        "[C-001] Business Registration Assistance",
        "[D-001] Revenue Distribution Management",
        "[E-001] Bank Account Activation Service",
        "[A-004] Payment Schedule Management",
        "[B-004] Compliance Report Generation",
    ],
}

# ══════════════════════════════════════════════════════════════════════
#  Card → Bill To lookup helpers
# ══════════════════════════════════════════════════════════════════════
# ── GitHub-API helpers for card_map persistence ─────────────────────────────
def _gh_cfg():
    """Return (token, repo, branch) from Streamlit secrets, or (None,None,None)."""
    try:
        cfg = st.secrets.get("github", {})
        return cfg.get("token"), cfg.get("repo"), cfg.get("branch", "main")
    except Exception:
        return None, None, None

def _gh_get(token, repo, branch, path="card_map.json"):
    """GET file from GitHub API. Returns (content_str, sha) or (None, None)."""
    import urllib.request, base64 as _b64
    url = f"https://api.github.com/repos/{repo}/contents/{path}?ref={branch}"
    req = urllib.request.Request(url, headers={
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json",
        "User-Agent": "igz-invoice-app"
    })
    try:
        with urllib.request.urlopen(req, timeout=8) as r:
            data = json.loads(r.read())
            content = _b64.b64decode(data["content"]).decode("utf-8")
            return content, data["sha"]
    except Exception:
        return None, None

def _gh_put(token, repo, branch, content_str, sha=None, path="card_map.json", msg="chore: update card_map"):
    """PUT (create or update) file on GitHub."""
    import urllib.request, base64 as _b64
    url = f"https://api.github.com/repos/{repo}/contents/{path}"
    body = {
        "message": msg,
        "content": _b64.b64encode(content_str.encode("utf-8")).decode("ascii"),
        "branch": branch,
    }
    if sha:
        body["sha"] = sha
    data = json.dumps(body).encode("utf-8")
    req = urllib.request.Request(url, data=data, method="PUT", headers={
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json",
        "Content-Type": "application/json",
        "User-Agent": "igz-invoice-app"
    })
    try:
        with urllib.request.urlopen(req, timeout=10):
            return True
    except Exception as e:
        st.warning(f"⚠️ GitHub sync failed: {e}")
        return False

def load_card_map() -> dict:
    """Load card_map: GitHub first, then local JSON fallback.
    Always merges CC_SEED so the 61 known cards are available
    even on a fresh deployment (user edits take precedence).
    """
    merged: dict = dict(CC_SEED)          # start with built-in seed
    token, repo, branch = _gh_cfg()
    if token and repo:
        content, _sha = _gh_get(token, repo, branch)
        if content:
            try:
                remote = json.loads(content)
                merged.update(remote)      # user edits override seed
                return merged
            except Exception:
                pass
    if CARD_MAP_FILE.exists():
        try:
            stored = json.loads(CARD_MAP_FILE.read_text(encoding="utf-8"))
            merged.update(stored)          # user edits override seed
        except Exception:
            pass
    return merged

def save_card_map(m: dict) -> None:
    """Save card_map to GitHub (auto-commit) + local fallback."""
    content_str = json.dumps(m, ensure_ascii=False, indent=2)
    try:
        CARD_MAP_FILE.write_text(content_str, encoding="utf-8")
    except Exception:
        pass
    token, repo, branch = _gh_cfg()
    if token and repo:
        _content, sha = _gh_get(token, repo, branch)
        ok = _gh_put(token, repo, branch, content_str, sha=sha,
                     msg="chore: update card_map via Streamlit UI")
        if ok:
            st.toast("💳 Card map synced to GitHub ✅", icon="✅")

def lookup_bill_to(card_no: str) -> str:
    """
    Given a masked card number (e.g. '4617 72** **** 3964'),
    extract last-4 digits and look up the cardholder name.
    Returns name string or '' if not found.
    """
    if not card_no:
        return ""
    # Extract last 4 consecutive digits
    digits = re.findall(r"[0-9]{4}", card_no.replace(" ", ""))
    last4 = digits[-1] if digits else ""
    if not last4:
        # Fallback: grab last 4 chars that are digits
        only_digits = re.sub(r"[^0-9]", "", card_no)
        last4 = only_digits[-4:] if len(only_digits) >= 4 else ""
    if not last4:
        return ""
    card_map = load_card_map()
    return card_map.get(last4, "")


def _packages(company: str) -> dict:
    """Return the package dict for the given company (uses same logic as _is_ast)."""
    return AST_PACKAGES if _is_ast(company) else IGZ_PACKAGES


# ══════════════════════════════════════════════════════════════════════
#  Auto-Select Items Algorithm  (v12 new)
# ══════════════════════════════════════════════════════════════════════
def auto_select_items(card_total: float, company: str,
                      min_items: int = 2, max_items: int = 5,
                      min_over_pct: float = 0.3,
                      max_over_pct: float = 25.0):
    """
    Find a combination of catalogue items whose sum EXCEEDS card_total by
    min_over_pct..max_over_pct percent.  Returns (selected_list, promo_rebate)
    where selected_list = [(name, price), ...] and promo_rebate < 0.
    Returns ([], 0) if no valid combo found.

    Algorithm:
      1. Filter products with 0 < price <= card_total
      2. Iterate combos from min_items to max_items
      3. Pick combo with smallest positive overshoot within allowed range
      4. If no combo found, relax max_over_pct to 40% and retry
    """
    import itertools
    products = _products(company)
    # Allow items up to 150% of card_total (enables single-item solutions for large amounts)
    eligible = [(name, price) for name, price in products if 0 < price <= card_total * 1.5]

    def _search(min_o, max_o):
        lo = card_total * (1 + min_o / 100)
        hi = card_total * (1 + max_o / 100)
        best = None
        best_score = float("inf")
        for n in range(min_items, max_items + 1):
            for combo in itertools.combinations(eligible, n):
                total = sum(p for _, p in combo)
                if lo <= total <= hi:
                    overshoot_pct = (total - card_total) / card_total * 100
                    score = overshoot_pct + n * 0.05   # prefer fewer items
                    if score < best_score:
                        best_score = score
                        best = combo
            if best:
                break   # stop at fewest items that satisfy constraint
        return best

    combo = _search(min_over_pct, max_over_pct)
    if combo is None:
        combo = _search(min_over_pct, 40.0)   # relaxed fallback

    # If still no combo, try combinations_with_replacement (allow repeated items)
    if combo is None:
        lo = card_total * (1 + min_over_pct / 100)
        hi = card_total * (1 + 40.0 / 100)
        best_rep = None
        best_rep_score = float("inf")
        for n in range(2, max_items + 2):
            for combo_r in itertools.combinations_with_replacement(eligible, n):
                total = sum(p for _, p in combo_r)
                if lo <= total <= hi:
                    overshoot_pct = (total - card_total) / card_total * 100
                    score = overshoot_pct + n * 0.05
                    if score < best_rep_score:
                        best_rep_score = score
                        best_rep = combo_r
            if best_rep:
                break
        combo = best_rep

    if combo:
        subtotal = sum(p for _, p in combo)
        promo    = -(subtotal - card_total)    # always negative
        return list(combo), round(promo, 2)
    return [], 0.0


# ══════════════════════════════════════════════════════════════════════
#  Excel column definitions
# ══════════════════════════════════════════════════════════════════════

# --- INFINITE GZ columns (v3: Items split into 4 cols) ---------------
IGZ_COLS = [
    "Invoice No",
    "Date",
    "Due Date",
    "Bill To",
    "Company",
    "Type",                # dropdown: VISA CREDIT/MASTERCARD/AMEXCARD/DEBIT/CASH/TRANSFER
    "Card No",
    "Approval Code",
    "Receipt No",
    "Ref No",
    # ↓ replaces "Items (JSON)"
    "Product Item",        # dropdown (IGZ transaction types)
    "Qty",
    "Unit Price (RM)",
    "Total Amount (RM)",
    # ↓ kept
    "Subtotal (RM)",
    "Promo Rebate (RM)",
    "Tax",
    "Total (RM)",
    "Remarks",
    "Receipt Link",        # wide + scroll
    "Invoice Link",        # wide + scroll
]

# --- AI SMART TECH columns (unchanged from v2) -----------------------
AST_COLS = [
    "Invoice No",
    "Date",
    "Due Date",
    "Bill To",
    "Company",
    "Type",
    "Card No",
    "Approval Code",
    "Receipt No",
    "Ref No",
    "Product Item",
    "Qty",
    "Unit Price (RM)",
    "Total Amount (RM)",
    "Subtotal (RM)",
    "Promo Rebate (RM)",
    "Tax",
    "Total (RM)",
    "Remarks",
    "Receipt Link",
    "Invoice Link",
]

def get_cols(company: str):
    return AST_COLS if _is_ast(company) else IGZ_COLS

def _is_ast(company: str) -> bool:
    """Exact-match check — avoids false positives like 'ASTRO'."""
    c = company.upper()
    return "AI SMART" in c or c == "AST" or c.startswith("AI SMART TECH")

def _products(company: str):
    return AST_PRODUCTS if _is_ast(company) else IGZ_PRODUCTS

# ══════════════════════════════════════════════════════════════════════
#  Excel helpers
# ══════════════════════════════════════════════════════════════════════
def _thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def init_excel(path: Path, company: str):
    """Load or create the Excel workbook, robust against corrupt/empty files."""
    import zipfile
    _need_create = True
    if path.exists() and path.stat().st_size > 100:
        try:
            wb = openpyxl.load_workbook(path)
            _need_create = False
        except (zipfile.BadZipFile, Exception):
            # Corrupt or empty file – wipe and recreate
            path.unlink(missing_ok=True)
    if _need_create:
        wb = openpyxl.Workbook()
        wb.active.title = "Invoices"
    # Ensure 'Invoices' sheet exists
    if "Invoices" not in wb.sheetnames:
        ws = wb.create_sheet("Invoices", 0)
        # Remove default 'Sheet' if present
        for sname in ["Sheet", "Sheet1"]:
            if sname in wb.sheetnames:
                del wb[sname]
    ws = wb["Invoices"]
    if ws.max_row == 0 or (ws.max_row == 1 and ws.cell(1,1).value is None):
        _write_header(ws, company)
        _add_dropdowns(ws, company)
        wb.save(path)
    return wb, ws


def _write_header(ws, company: str):
    COLS = get_cols(company)
    if _is_ast(company):
        hdr_fill = PatternFill("solid", fgColor="6B5B95")
    else:
        hdr_fill = PatternFill("solid", fgColor="1A1A1A")
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    # Row 1: Banner
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    c = ws["A1"]
    c.value     = f"{company}  —  Invoice Records"
    c.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    c.fill      = hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Row 2: Column headers
    COL_W = _col_widths(company)
    for ci, col in enumerate(COLS, 1):
        cell = ws.cell(row=2, column=ci, value=col)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin()
        ws.column_dimensions[get_column_letter(ci)].width = COL_W.get(col, 14)
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "A3"


def _col_widths(company: str) -> dict:
    # Both companies now have the same 4-column product structure
    base = {
        "Invoice No": 13, "Date": 14, "Due Date": 14, "Bill To": 16,
        "Company": 22, "Type": 18, "Card No": 22,
        "Approval Code": 14, "Receipt No": 12, "Ref No": 18,
        "Product Item": 48,
        "Qty": 6, "Unit Price (RM)": 15, "Total Amount (RM)": 16,
        "Subtotal (RM)": 14, "Promo Rebate (RM)": 17,
        "Tax": 7, "Total (RM)": 13,
        "Remarks": 30,
        "Receipt Link": 42,
        "Invoice Link": 42,
    }
    return base


def _add_dropdowns(ws, company: str):
    """Add Excel DataValidation dropdowns for all companies."""
    COLS = get_cols(company)
    is_ast = _is_ast(company)

    def col_letter(name):
        idx = COLS.index(name) + 1
        return get_column_letter(idx)

    # ── Type dropdown (merged Payment Type + Card Type) ─────────────
    if "Type" in COLS:
        dv_type = DataValidation(
            type="list",
            formula1='"VISA CREDIT,MASTERCARD,AMEXCARD,DEBIT,CASH,TRANSFER"',
            allow_blank=True, showDropDown=False,
        )
        dv_type.sqref       = f"{col_letter('Type')}3:{col_letter('Type')}500"
        dv_type.prompt      = "请选择付款类型"
        dv_type.promptTitle = "Type"
        ws.add_data_validation(dv_type)

    # ── Product Item dropdown (both companies, different lists) ──
    if "Product Item" in COLS:
        products = _products(company)
        sheet_name = "IGZProductList" if not is_ast else "ASTProductList"
        try:
            if sheet_name not in [s.title for s in ws.parent.worksheets]:
                pl_ws = ws.parent.create_sheet(sheet_name)
            else:
                pl_ws = ws.parent[sheet_name]

            for i, (label, price) in enumerate(products, 1):
                pl_ws.cell(row=i, column=1, value=label)
                pl_ws.cell(row=i, column=2, value=price)

            max_row = len(products)
            dv_prod = DataValidation(
                type="list",
                formula1=f"{sheet_name}!$A$1:$A${max_row}",
                allow_blank=True, showDropDown=False,
            )
            dv_prod.sqref       = f"{col_letter('Product Item')}3:{col_letter('Product Item')}500"
            dv_prod.prompt      = "从下拉选择服务/交易类型"
            dv_prod.promptTitle = "Product Item"
            ws.add_data_validation(dv_prod)
            pl_ws.sheet_state = "hidden"
        except Exception as e:
            print(f"[WARN] Product Item dropdown: {e}")


def append_row(wb, ws, data: dict, path: Path, company: str):
    COLS = get_cols(company)
    is_ast = _is_ast(company)
    even_fill = PatternFill("solid", fgColor="F9F5FF" if is_ast else "FFF8F0")
    odd_fill  = PatternFill("solid", fgColor="FFFFFF")
    lc = "6B5B95" if is_ast else "1A1A1A"

    row_num = ws.max_row + 1
    fill    = even_fill if row_num % 2 == 0 else odd_fill

    for ci, col in enumerate(COLS, 1):
        val  = data.get(col, "")
        cell = ws.cell(row=row_num, column=ci, value=val)
        cell.fill      = fill
        cell.alignment = Alignment(vertical="center", wrap_text=(col == "Product Item"))
        cell.border    = _thin()
        if col in ("Receipt Link", "Invoice Link") and val:
            cell.font = Font(name="Calibri", color=lc, underline="single", size=9)
        else:
            cell.font = Font(name="Calibri", size=9)

    ws.row_dimensions[row_num].height = 20
    wb.save(path)
    return row_num


def update_cell(wb, ws, row_num: int, col_name: str, value, path: Path, company: str):
    COLS = get_cols(company)
    if col_name not in COLS:
        return
    ci   = COLS.index(col_name) + 1
    cell = ws.cell(row=row_num, column=ci, value=value)
    lc   = "6B5B95" if _is_ast(company) else "1A1A1A"
    if col_name in ("Receipt Link", "Invoice Link") and value:
        cell.font = Font(name="Calibri", color=lc, underline="single", size=9)
    wb.save(path)


# ══════════════════════════════════════════════════════════════════════
#  OCR + parser
# ══════════════════════════════════════════════════════════════════════
def ocr_image(img: Image.Image) -> str:
    return pytesseract.image_to_string(img, lang="eng")


def _clean_inv_no(raw: str) -> str:
    """Strip INV- / INV_ prefix; return bare number/code string."""
    s = raw.strip()
    s = re.sub(r"(?i)^INV[-_\s]*", "", s)
    return s


def parse_receipt(text: str, company: str) -> dict:
    d = {}
    def _s(patterns):
        for pat in patterns:
            m = re.search(pat, text, re.IGNORECASE | re.MULTILINE)
            if m:
                return m.group(1).strip()
        return ""

    # ── Invoice No ────────────────────────────────────────────────
    # Matches: INV NO 000962 / INVOICE NO / #000962 / NO. 000962
    # NOTE: POS card receipts typically do NOT have invoice numbers.
    # Invoice No is populated if the receipt is a tax invoice / OR
    # the user manually types it in Step 2.
    raw_inv = _s([
        r"invoice\s*no[.:\s]+([0-9A-Z\-_/]+)",
        r"[il]nvoice\s*no\.?\s*[:\s]+([0-9A-Z\-_/]+)",
        r"inv(?:oice)?\s*no[.:\s]+([A-Z0-9_\-/]+)",
        r"inv[-_]?no[.:\s]+([A-Z0-9_\-/]+)",
        r"inv\s*no[.:\s]+([0-9]{3,})",
        r"#\s*([0-9]{4,})",                    # #000962
        r"no\.\s*([0-9]{4,})",               # NO. 000962
        r"receipt\s*#\s*([0-9A-Z\-]{4,})",  # RECEIPT # 2026-001
    ])
    d["Invoice No"] = _clean_inv_no(raw_inv) if raw_inv else ""

    # ── Receipt No / Trace No ─────────────────────────────────────
    # Covers: RECEIPT NO / TRACE NO. / RETRIEVAL REF NO / STAN / RRN / SEQ NO
    # Also handles dots in labels: TRACE NO. / RETRIEVAL REF. NO.
    d["Receipt No"] = _s([
        r"receipt\s*no\.?\s*:\s*trace\s*no\.?\s*[\n\r]+\s*([0-9]+)",  # multiline HLB
        r"receipt\s*no[.:\s]+([0-9]{3,})",
        r"trace\s*no\.?\s*[.:\s]+([0-9]{3,})",                  # TRACE NO. / TRACE NO:
        r"trace\s*number\s*[:\s]+([0-9]{3,})",
        r"stan\s*/\s*trace[.:\s]+([0-9]{3,})",
        r"stan[.:\s]+([0-9]{3,})",
        r"rrn\s*[:\s]+([0-9]{6,})",
        r"seq\s*no[.:\s]+([0-9]{3,})",
        r"sequence\s*no[.:\s]+([0-9]{3,})",
        r"host\s*trace\s*[:\s]+([0-9]{3,})",
        r"trace[:\s]+([0-9]{3,})",
    ])

    # ── Date ──────────────────────────────────────────────────────
    # Handles: 26MAR2026 / 26/03/2026 / 26-03-2026 / 26 MAR 2026
    raw_date = _s([
        r"date[/\\]?time[\s.:]+([0-9]{1,2}[A-Z]{3}[0-9]{4})",
        r"date[\s.:]+([0-9]{1,2}[A-Z]{3}[0-9]{4})",
        r"date[\s.:]+([0-9]{1,2}[\s/\-][0-9]{1,2}[\s/\-][0-9]{2,4})",
        r"([0-9]{1,2}[A-Z]{3}[0-9]{4})",
        r"([0-9]{2}/[0-9]{2}/[0-9]{4})",
    ])
    # Convert 26MAR2026 → 26/03/2026
    if raw_date:
        month_map = {"JAN":"01","FEB":"02","MAR":"03","APR":"04",
                     "MAY":"05","JUN":"06","JUL":"07","AUG":"08",
                     "SEP":"09","OCT":"10","NOV":"11","DEC":"12"}
        m2 = re.match(r"([0-9]{1,2})([A-Z]{3})([0-9]{4})", raw_date.upper())
        if m2:
            dd, mon, yyyy = m2.group(1), m2.group(2), m2.group(3)
            mm = month_map.get(mon, mon)
            raw_date = f"{dd.zfill(2)}/{mm}/{yyyy}"
    d["Date"] = raw_date or ""
    d["Due Date"] = d["Date"]

    # ── Time ──────────────────────────────────────────────────────
    d["Time"] = _s([
        r"date[/\\]?time\s+[0-9A-Z]+\s+([0-9]{1,2}:[0-9]{2}:[0-9]{2})",
        r"time[\s.:]+([0-9]{1,2}:[0-9]{2}(?::[0-9]{2})?)",
        r"[0-9]{1,2}[A-Z]{3}[0-9]{4}\s+([0-9]{2}:[0-9]{2}:[0-9]{2})",
    ])

    # ── Bill To (merchant name on receipt = payee) ────────────────
    # Covers: BILL TO / ISSUED TO / CARDHOLDER NAME / NAME: / CARD HOLDER
    # Also handles Maybank-style (no label – grab caps after card no line)
    raw_bill = _s([
        r"bill\s*to[\s.:]+([A-Za-z0-9 .,'&/-]+?)(?=\n|company|payment|\Z)",
        r"issued\s*to[\s.:]+([A-Za-z0-9 .,'&/-]+?)(?=\n|company|\Z)",
        r"cardholder\s*name[\s.:]+([A-Za-z0-9 .,'&/-]+?)(?=\n|\Z)",
        r"card\s*holder\s*name[\s.:]+([A-Za-z0-9 .,'&/-]+?)(?=\n|\Z)",
        r"card\s*holder[\s.:]+([A-Za-z0-9 .,'&/-]+?)(?=\n|\Z)",
        r"name[\s.:]+([A-Z][A-Za-z0-9 .,'&/-]{3,})(?=\n|\Z)",
        # NOTE: all-caps catch-all removed (grabs bank names like PUBLIC BANK)  # CC_SEED/card_map is the authoritative Bill To source
    ])
    d["Bill To"] = re.sub(r"\s+", " ", (raw_bill or "")).strip()

    # ── Auto-lookup Bill To from card map (ALWAYS override OCR grab) ────────
    # card_map is authoritative: even if OCR wrongly filled Bill To with a
    # bank name (e.g. "PUBLIC BANK"), the card_map lookup will correct it.
    if d.get("Card No"):
        _auto_bt = lookup_bill_to(d["Card No"])
        if _auto_bt:
            d["Bill To"] = _auto_bt
            d["_bill_to_source"] = "card_map"   # flag for UI badge

    # ── Type (merged: Payment Network + Credit/Debit) ────────────
    # Receipt line: "VISA CREDIT" / "MASTERCARD" / "AMEX" / "DEBIT" / "CASH" / "TRANSFER"
    # Maps to dropdown: VISA CREDIT, MASTERCARD, AMEXCARD, DEBIT, CASH, TRANSFER
    _type_raw = _s([
        r"(visa\s+credit)",           # VISA CREDIT  (exact receipt line)
        r"(mastercard\s+credit)",     # MASTERCARD CREDIT
        r"(master\s+credit)",
        r"(mastercard)",               # bare MASTERCARD
        r"(amex\s+credit)",
        r"(amexcard)",
        r"(amex)",
        r"host\s+(visa|mastercard|master|amex)",   # HOST VISA line
        r"\b(debit)\b",
        r"\b(cash)\b",
        r"\b(transfer)\b",
    ])
    _type_map = {
        "visa credit":       "VISA CREDIT",
        "mastercard credit": "MASTERCARD",
        "master credit":     "MASTERCARD",
        "mastercard":        "MASTERCARD",
        "amex credit":       "AMEXCARD",
        "amexcard":          "AMEXCARD",
        "amex":              "AMEXCARD",
        "visa":              "VISA CREDIT",
        "master":            "MASTERCARD",
        "debit":             "DEBIT",
        "cash":              "CASH",
        "transfer":          "TRANSFER",
    }
    _type_key = (_type_raw or "").lower().strip()
    d["Type"] = _type_map.get(_type_key, (_type_raw or "").upper())

    # Keep legacy keys for PDF builder compatibility (will be removed in v13)
    if "VISA" in d["Type"]:
        d["Payment Type"] = "Visa"; d["Card Type"] = "CREDIT"
    elif "MASTERCARD" in d["Type"]:
        d["Payment Type"] = "Master"; d["Card Type"] = "CREDIT"
    elif "AMEX" in d["Type"]:
        d["Payment Type"] = "Amex"; d["Card Type"] = "CREDIT"
    elif "DEBIT" in d["Type"]:
        d["Payment Type"] = ""; d["Card Type"] = "DEBIT"
    else:
        d["Payment Type"] = d["Type"]; d["Card Type"] = ""

    # ── Card No ──────────────────────────────────────────────────
    # Matches many Malaysian POS formats:
    #   4617 72** **** 3964   (asterisk mask)
    #   4617 72XX XXXX 3964   (X mask, HLB/Maybank)
    #   4617 72xx xxxx 3964   (lowercase x)
    #   4617 XXXX XXXX 3964   (fully masked middle)
    #   CARD : 4617 72XX XXXX 3964
    # Masked char set: OCR may read mask as * X x . • – or digit
    _MASK = r"[0-9X*x.•\-]"
    _card_raw = _s([
        # With "CARD NO:" label (highest confidence)
        r"card\s*no\.?[\s.:]+([0-9]{4}[\s-]" + _MASK + r"{2,6}[\s-]" + _MASK + r"{2,6}[\s-][0-9]{4})",
        r"card[:\s]+([0-9]{4}[\s-]" + _MASK + r"{2,6}[\s-]" + _MASK + r"{2,6}[\s-][0-9]{4})",
        # Bare 4-group format on its own line (no label) — broadened mask chars
        r"([0-9]{4}\s+" + _MASK + r"{2,6}\s+" + _MASK + r"+\s+[0-9]{4})",
        # Full-star mask: ************8888
        r"card\s*no\.?[\s.:]+([*]{8,}\s*[0-9]{4})",
        r"([*]{8,}[0-9]{4})",
        # Looser catch-all with card no label
        r"card\s*no\.?[\s.:]+([0-9*Xx\s]{13,25})",
    ])
    # Normalise: replace lowercase x → X for consistency
    d["Card No"] = re.sub(r"x", "X", _card_raw).strip() if _card_raw else ""

    # ── Fallback: if Card No still empty, try last-4 from any partial card line ─
    if not d["Card No"]:
        # Matches lines like "4617 72** **** 3964" even with unusual OCR chars
        m_last4 = re.search(
            r"([0-9]{4}[\s\S]{4,20}?\b([0-9]{4})\b)(?:\s|$)",
            text, re.IGNORECASE | re.MULTILINE
        )
        if m_last4:
            d["Card No"] = m_last4.group(1).strip()

    # ── Approval Code ────────────────────────────────────────────
    # Pad to 6 chars with leading zeros (some banks strip leading zero, e.g. 85944 → 085944)
    _appr_raw = _s([
        r"approval\s*code\s*[:\s]+([A-Z0-9]+)",
        r"approval\s*code[\s.:]+([A-Z0-9]+)",
        r"auth(?:orisation)?\s*code[\s.:]+([A-Z0-9]+)",
        r"auth\s*no\.?[\s.:]+([A-Z0-9]+)",
        r"approval[\s.:]+([A-Z0-9]+)",
    ])
    # If purely numeric and shorter than 6 digits, left-pad with zeros
    if _appr_raw and _appr_raw.isdigit() and len(_appr_raw) < 6:
        _appr_raw = _appr_raw.zfill(6)
    d["Approval Code"] = _appr_raw or ""

    # ── Ref No ───────────────────────────────────────────────────
    # Covers: REF NO / REFERENCE NO / RETRIEVAL REF / RETRIEVAL REF. NO. / HOST REF
    d["Ref No"] = _s([
        r"retrieval\s*ref\.?\s*no\.?\s*[:\s]+([0-9A-Z]{6,})",  # RETRIEVAL REF. NO.
        r"retrieval\s*ref(?:erence)?[:\s]+([0-9A-Z]{6,})",
        r"ref(?:erence)?\s*no[.:\s]+([0-9A-Z]{6,})",
        r"host\s*ref(?:erence)?[:\s]+([0-9A-Z]{6,})",
        r"terminal\s*ref[:\s]+([0-9A-Z]{6,})",
        r"ref\.?\s*no\.?\s*[:\s]+([0-9A-Z]{6,})",               # REF. NO. (with dots)
        r"\bref\b[:\s]+([0-9]{6,})",
        r"ref\s*no[.:\s]+([0-9]{4,})",
    ])

    # ── Total ────────────────────────────────────────────────────
    # Covers: TOTAL RM 18888 / AMOUNT RM / RM 18888.00 (bare, last numeric)
    total_s = _s([
        r"total\s+rm\s*([\d,]+\.?\d*)",
        r"total[\s.:]+rm\s*([\d,]+\.?\d*)",
        r"amount[\s.:]+rm\s*([\d,]+\.?\d*)",
        r"(?:grand\s*)?total[:\s]+([\d,]+\.\d{2})",
        r"^rm\s*([\d,]+\.\d{2})\s*$",   # bare "RM 18888.00" on its own line
        r"rm\s*([1-9][\d,]{2,}\.\d{2})",  # RM followed by significant amount
    ])
    d["Total (RM)"] = float(total_s.replace(",","")) if total_s else ""

    # ── Subtotal ─────────────────────────────────────────────────
    sub_s = _s([r"subtotal[\s.:]+rm\s*([\d,]+\.?\d*)"])
    d["Subtotal (RM)"] = float(sub_s.replace(",","")) if sub_s else d.get("Total (RM)","")

    # ── Promo / Discount ─────────────────────────────────────────
    promo_s = _s([
        r"promo\s*rebate[\s.:(]+rm\s*([\d,]+\.?\d*)",
        r"discount[\s.:(]+rm\s*([\d,]+\.?\d*)",
    ])
    d["Promo Rebate (RM)"] = f"-{promo_s}" if promo_s else ""

    # ── Tax / Remarks ─────────────────────────────────────────────
    # Restrict tax to numeric values only (avoids capturing "TAX INVOICE 2026-001")
    d["Tax"]     = _s([r"tax[\s.:]+([\d][\d\.\-]*\s*%?)",
                       r"service\s*tax[\s.:]+([\d][\d\.\-]*)",
                       r"gst[\s.:]+([\d][\d\.\-]*)",
                       ]) or "-"
    d["Remarks"] = _s([r"remarks?[\s.:]+([^\n]+)"]) or ""

    # ── Product item defaults ─────────────────────────────────────
    d["Product Item"]      = ""
    d["Qty"]               = 1
    d["Unit Price (RM)"]   = d.get("Total (RM)", "")
    d["Total Amount (RM)"] = d.get("Total (RM)", "")
    d["Company"]           = ""   # always initialise; user fills in Step 2

    return d


# ══════════════════════════════════════════════════════════════════════
#  Invoice PDF generator wrapper
# ══════════════════════════════════════════════════════════════════════
def generate_invoice_pdf(company: str, data: dict, inv_number: str,
                         items_list: list = None) -> Path:
    items = items_list or []
    if not items:
        try:
            unit_p = float(str(data.get("Unit Price (RM)","") or data.get("Total (RM)","0")).replace(",",""))
            qty    = int(data.get("Qty", 1) or 1)
            amt    = float(str(data.get("Total Amount (RM)","") or data.get("Total (RM)","0")).replace(",",""))
            desc   = data.get("Product Item","Services rendered") or "Services rendered"
            # Strip code prefix [X-000] for cleaner PDF display
            desc_clean = re.sub(r"^\[[A-Z]-\d{3}\]\s*", "", desc)
            items  = [{"desc": desc_clean, "unit_price": unit_p, "qty": qty, "amount": amt}]
        except Exception:
            total = float(str(data.get("Total (RM)", 0) or 0).replace(",",""))
            items = [{"desc": "Services rendered", "unit_price": total, "qty": 1, "amount": total}]

    inv_no_clean = _clean_inv_no(str(inv_number))

    inv = {
        "inv_no":       inv_no_clean,
        "date":         data.get("Date",""),
        "due_date":     data.get("Due Date", data.get("Date","")),
        "bill_to":      data.get("Bill To",""),
        "company":      data.get("Company","–"),
        "payment_type": data.get("Type", data.get("Payment Type","")),
        "card_type":    data.get("Card Type",""),  # kept for PDF compat
        "card_no":      data.get("Card No",""),
        "approval":     data.get("Approval Code",""),
        "receipt_no":   data.get("Receipt No",""),
        "ref_no":       data.get("Ref No",""),
        "tax":          data.get("Tax","-"),
        "total":        float(str(data.get("Total (RM)",0) or 0).replace(",","")),
        "items":        items,
        "remarks":      [data.get("Remarks","")] if data.get("Remarks") else [],
    }
    promo = data.get("Promo Rebate (RM)","")
    if promo:
        try:
            # Always store promo as NEGATIVE (deduction)
            _pv = abs(float(str(promo).replace(",","").lstrip("-")))
            inv["promo_rebate"] = -_pv
        except: pass

    ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_no  = re.sub(r"[^A-Za-z0-9_\-]","_", inv_no_clean)
    safe_no  = safe_no or ts   # fallback if inv_no_clean is empty
    pdf_path = INVOICES_DIR / f"invoice_{safe_no}_{ts}.pdf"

    if _is_ast(company):
        make_ast(inv, str(pdf_path))
    else:
        make_igz(inv, str(pdf_path))
    return pdf_path


# ══════════════════════════════════════════════════════════════════════
#  Streamlit UI
# ══════════════════════════════════════════════════════════════════════
def main():
    st.set_page_config(
        page_title="Invoice Management System",
        page_icon="🧾",
        layout="wide",
        initial_sidebar_state="collapsed",
    )

    st.markdown("""
    <style>
    .stApp { background: #F7F8FA; }
    div[data-testid="stDataFrame"] > div { overflow-x: auto !important; }
    div[data-testid="stDataFrame"] iframe { width: 100% !important; }
    .section-title { font-size:17px; font-weight:700; margin:4px 0 10px 0; color:#1A1A1A; }
    .info-box { background:#fff; border-radius:10px; padding:14px 18px;
                border:1px solid #E5E5E5; margin-bottom:12px; }
    .price-row { font-size:11px; padding:2px 0; border-bottom:1px solid #eee; }
    </style>
    """, unsafe_allow_html=True)

    # ── Session state ─────────────────────────────────────────────
    for k, v in [("company",None),("parsed_data",{}),("receipt_path",None),
                 ("excel_row",None),("invoice_done",False),("invoice_path",None),
                 ("items_list",[]),
                 ("item_select_mode","🤖 自动选品（Auto-Select）"),
                 ("auto_suggest_items",[]),  ("auto_suggest_promo",0.0),
                 ("auto_suggest_cache_key",""), ("auto_items_for_pdf",[])]:
        if k not in st.session_state:
            st.session_state[k] = v

    # ══════════════════════════════════════════════════════════════
    #  SCREEN 1 — Company selection
    # ══════════════════════════════════════════════════════════════
    if st.session_state.company is None:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("## 🧾 Invoice Management System")
        st.markdown("请选择公司 / Select company to proceed:")
        st.markdown("<br>", unsafe_allow_html=True)

        c1, c2, c3 = st.columns([1, 1, 1.2])

        # ── IGZ card ──────────────────────────────────────────────
        with c1:
            st.markdown("""
            <div style='background:linear-gradient(135deg,#1A1A1A,#333);color:#fff;
                        border-radius:14px;padding:28px 20px;text-align:center;'>
                <div style='font-size:34px'>🏢</div>
                <div style='font-size:19px;font-weight:800;margin:8px 0'>INFINITE GZ SDN BHD</div>
                <div style='font-size:11px;opacity:.7'>202401019141 (1564990-X)</div>
                <div style='font-size:11px;opacity:.6;margin-top:4px'>Hong Leong Bank · Black Minimal Style</div>
                <div style='font-size:10px;opacity:.5;margin-top:4px'>51 Transaction Types | 5 Groups</div>
            </div>""", unsafe_allow_html=True)
            if st.button("选择 INFINITE GZ  →", key="btn_igz", use_container_width=True):
                st.session_state.company = "INFINITE GZ SDN BHD"; st.rerun()

        # ── AST card ──────────────────────────────────────────────
        with c2:
            st.markdown("""
            <div style='background:linear-gradient(135deg,#6B5B95,#9B8EC4);color:#fff;
                        border-radius:14px;padding:28px 20px;text-align:center;'>
                <div style='font-size:34px'>💼</div>
                <div style='font-size:19px;font-weight:800;margin:8px 0'>AI SMART TECH SDN BHD</div>
                <div style='font-size:11px;opacity:.7'>202401043356 (1589202-V)</div>
                <div style='font-size:11px;opacity:.6;margin-top:4px'>Purple Corporate · AI Tech Services</div>
                <div style='font-size:10px;opacity:.5;margin-top:4px'>37 AI Services | 2026 Market Prices</div>
            </div>""", unsafe_allow_html=True)
            if st.button("选择 AI SMART TECH →", key="btn_ast", use_container_width=True):
                st.session_state.company = "AI SMART TECH SDN BHD"; st.rerun()

        # ── RIGHT COLUMN: both price tables ──────────────────────
        with c3:
            tab_igz, tab_ast = st.tabs(["🏢 IGZ 服务价格", "💼 AST 服务价格"])

            with tab_igz:
                st.markdown("**INFINITE GZ — 2026 业务服务参考价**")
                # Group headers
                groups = {
                    "Group A · 融资安排": [(l,p) for l,p in IGZ_PRODUCTS if l.startswith("[A-")],
                    "Group B · 记录处理": [(l,p) for l,p in IGZ_PRODUCTS if l.startswith("[B-")],
                    "Group C · 数字服务": [(l,p) for l,p in IGZ_PRODUCTS if l.startswith("[C-")],
                    "Group D · 成果分享": [(l,p) for l,p in IGZ_PRODUCTS if l.startswith("[D-")],
                    "Group E · 账户操作": [(l,p) for l,p in IGZ_PRODUCTS if l.startswith("[E-")],
                }
                for gname, items in groups.items():
                    st.markdown(f"<div style='font-size:10px;font-weight:700;color:#555;margin-top:6px'>{gname}</div>",
                                unsafe_allow_html=True)
                    for label, price in items[:4]:
                        short = re.sub(r"^\[[A-Z]-\d{3}\]\s*","",label)
                        price_str = f"RM {price:,}" if price > 0 else "–"
                        st.markdown(
                            f"<div class='price-row'><b>{price_str}</b> · {short[:38]}</div>",
                            unsafe_allow_html=True)
                st.caption("完整列表见 Product Item 下拉菜单（51项）")

            with tab_ast:
                st.markdown("**AI SMART TECH — 2026 AI 服务参考价**")
                for label, price in AST_PRODUCTS[:14]:
                    st.markdown(
                        f"<div class='price-row'><b>RM {price:,}</b> · {label[:42]}</div>",
                        unsafe_allow_html=True)
                st.caption("完整列表见 Product Item 下拉菜单（37项）")

        return


    # ── Sidebar: Card Map Manager (always visible after company selected) ──
    with st.sidebar:
        st.markdown("### 📇 卡号对照表")
        st.caption("卡尾4位 → 自动识别 Bill To（持卡人姓名）")
        _cmap = load_card_map()

        # ── Show existing entries ──────────────────────────────────
        if _cmap:
            st.markdown("**已登记的卡号：**")
            _to_delete = None
            for _l4, _nm in sorted(_cmap.items()):
                _col1, _col2 = st.columns([3, 1])
                _col1.markdown(f"**`**** {_l4}`** → {_nm}")
                if _col2.button("🗑️", key=f"del_{_l4}", help=f"删除 {_l4}"):
                    _to_delete = _l4
            if _to_delete:
                del _cmap[_to_delete]
                save_card_map(_cmap)
                st.success(f"已删除卡尾 {_to_delete}")
                st.rerun()
        else:
            st.info("对照表为空，请添加第一张卡")

        st.markdown("---")
        st.markdown("**➕ 添加新卡**")
        _new_l4   = st.text_input("卡尾4位数字", max_chars=4, key="new_card_last4",
                                   placeholder="例：3964")
        _new_name = st.text_input("持卡人姓名", key="new_card_name",
                                   placeholder="例：Max Lai")
        if st.button("✅ 保存", key="save_card_map"):
            _new_l4 = re.sub(r"[^0-9]", "", _new_l4.strip())
            _new_name = _new_name.strip()
            if len(_new_l4) == 4 and _new_name:
                _cmap[_new_l4] = _new_name
                save_card_map(_cmap)
                st.success(f"✅ 已保存：**** {_new_l4} → {_new_name}")
                st.rerun()
            else:
                st.error("请输入正确的4位卡尾数字 + 持卡人姓名")

        st.markdown("---")
        st.caption("📌 对照表储存于 output/card_map.json\n下次启动自动载入")

    # ══════════════════════════════════════════════════════════════
    #  SCREEN 2 — Main workflow
    # ══════════════════════════════════════════════════════════════
    company = st.session_state.company
    is_ast  = _is_ast(company)
    bar_col = "#6B5B95" if is_ast else "#1A1A1A"

    st.markdown(f"""
    <div style='background:{bar_col};color:#fff;padding:13px 20px;border-radius:10px;
                display:flex;align-items:center;justify-content:space-between;margin-bottom:16px'>
        <span style='font-size:19px;font-weight:800'>{'💼' if is_ast else '🏢'} {company}</span>
        <span style='font-size:11px;opacity:.7'>Invoice Management System v3 (OCR v12.5)</span>
    </div>""", unsafe_allow_html=True)

    if st.button("← 切换公司 / Switch Company"):
        for k in ["company","parsed_data","receipt_path","excel_row",
                  "invoice_done","invoice_path","items_list"]:
            st.session_state[k] = {} if k in ("parsed_data",) else ([] if k=="items_list" else None)
        st.session_state.invoice_done = False
        # Also reset package selection
        st.session_state["selected_package"] = ""
        st.session_state["pkg_sub_items"] = []
        st.rerun()

    excel_path = AST_EXCEL if is_ast else IGZ_EXCEL
    wb, ws = init_excel(excel_path, company)

    # ── STEP 1: Upload receipt ─────────────────────────────────────
    st.markdown("---")
    st.markdown("<div class='section-title'>📤 Step 1 · 上传收据</div>", unsafe_allow_html=True)
    uploaded = st.file_uploader("上传收据图片 (PNG / JPG)", type=["png","jpg","jpeg","webp"])

    if uploaded and st.session_state.receipt_path is None:
        ts  = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        ext = Path(uploaded.name).suffix or ".png"
        rec_path = RECEIPTS_DIR / f"receipt_{ts}{ext}"
        with open(rec_path,"wb") as f: f.write(uploaded.read())
        st.session_state.receipt_path = str(rec_path)

        try:
            with st.spinner("🔍 正在 OCR 解析收据…"):
                img    = Image.open(rec_path)
                raw    = ocr_image(img)
                parsed = parse_receipt(raw, company)
                parsed["Receipt Link"] = str(rec_path)
                st.session_state.parsed_data = parsed
                row_num = append_row(wb, ws, parsed, excel_path, company)
                st.session_state.excel_row = row_num
            st.success(f"✅ 收据已解析，已写入 Excel 第 {row_num} 行")
            st.rerun()
        except Exception as _ocr_err:
            st.error(f"❌ OCR 解析失败: {_ocr_err}")
            st.session_state.receipt_path = None  # allow re-upload

    if st.session_state.receipt_path:
        with st.expander("📎 已上传的收据", expanded=False):
            try: st.image(st.session_state.receipt_path, width=420)
            except: st.write(st.session_state.receipt_path)

    # ── STEP 2: Review & Edit ─────────────────────────────────────
    if st.session_state.parsed_data:
        st.markdown("---")
        st.markdown("<div class='section-title'>✏️ Step 2 · 确认 / 编辑内容</div>",
                    unsafe_allow_html=True)
        _bt_from_map = st.session_state.parsed_data.get("_bill_to_source") == "card_map"
        if _bt_from_map:
            st.success("✅ 系统已自动解析收据，并从卡号对照表自动识别 **Bill To**。\n「Company」= 客户公司名称（无则填 -）")
        else:
            st.info("📌 系统已自动解析收据。\n① 「Bill To」= 卡号未在对照表中 → 请手动填写客户姓名\n② 「Company」= 客户公司名称（无则填 -）\n📇 可在右侧侧边栏「卡号对照表」添加新卡")

        d = st.session_state.parsed_data
        c1, c2 = st.columns(2)

        # ── Left column ──────────────────────────────────────────
        with c1:
            d["Invoice No"]  = st.text_input(
                "Invoice No（收据实际编号）",
                d.get("Invoice No",""),
                help="系统自动去除 INV- 前缀，保留收据原始编号")
            d["Date"]        = st.text_input("Date", d.get("Date",""))
            d["Due Date"]    = st.text_input("Due Date", d.get("Due Date",""))
            _bt_label = (
                "Bill To 💳 (卡号对照表自动填入 — 可修改)"
                if st.session_state.parsed_data.get("_bill_to_source") == "card_map"
                else "Bill To (持卡人/客户姓名)"
            )
            d["Bill To"] = st.text_input(_bt_label, d.get("Bill To",""))
            d["Company"]     = st.text_input(
                "⭐ Company (手动填写，无则填 -)",
                d.get("Company",""),
                help="唯一需手动输入的字段")

            # ── TYPE dropdown (merged Payment + Card type) ────────────
            type_opts = ["","VISA CREDIT","MASTERCARD","AMEXCARD","DEBIT","CASH","TRANSFER"]
            cur_type  = d.get("Type","")
            # backward-compat: if old Payment Type existed, derive Type
            if not cur_type and d.get("Payment Type"):
                _pt = d.get("Payment Type","").lower()
                _ct = d.get("Card Type","").lower()
                if "visa"   in _pt: cur_type = "VISA CREDIT"
                elif "master" in _pt: cur_type = "MASTERCARD"
                elif "amex"   in _pt: cur_type = "AMEXCARD"
                elif "debit"  in _ct: cur_type = "DEBIT"
                elif "cash"   in _pt: cur_type = "CASH"
                elif "transfer" in _pt: cur_type = "TRANSFER"
            idx_type  = type_opts.index(cur_type) if cur_type in type_opts else 0
            d["Type"] = st.selectbox("💳 Type", type_opts, index=idx_type,
                                      help="VISA CREDIT / MASTERCARD / AMEXCARD / DEBIT / CASH / TRANSFER")
            # Sync legacy fields for PDF builder
            if "VISA"   in d["Type"]: d["Payment Type"]="Visa";   d["Card Type"]="CREDIT"
            elif "MASTER" in d["Type"]: d["Payment Type"]="Master"; d["Card Type"]="CREDIT"
            elif "AMEX"   in d["Type"]: d["Payment Type"]="Amex";   d["Card Type"]="CREDIT"
            elif "DEBIT"  in d["Type"]: d["Payment Type"]="";       d["Card Type"]="DEBIT"
            else: d["Payment Type"]=d["Type"]; d["Card Type"]=""

        # ── Right column ─────────────────────────────────────────
        with c2:
            d["Card No"]       = st.text_input("Card No",       d.get("Card No",""))
            d["Approval Code"] = st.text_input("Approval Code", d.get("Approval Code",""))
            d["Receipt No"]    = st.text_input("Receipt No",    d.get("Receipt No",""))
            d["Ref No"]        = st.text_input("Ref No",        d.get("Ref No",""))
            d["Tax"]           = st.text_input("Tax",           d.get("Tax","-"))
            d["Subtotal (RM)"] = st.text_input("Subtotal (RM)", str(d.get("Subtotal (RM)","")))
            d["Promo Rebate (RM)"] = st.text_input("Promo Rebate (RM)",
                                                    str(d.get("Promo Rebate (RM)","")))
            # ── Discount guard: warn if promo rebate > 15% of subtotal ──
            try:
                _sub_val   = float(str(d.get("Subtotal (RM)","0")).replace(",",""))
                _promo_raw = str(d.get("Promo Rebate (RM)","")).lstrip("-").replace(",","")
                _promo_val = float(_promo_raw) if _promo_raw else 0.0
                if _sub_val > 0 and _promo_val > 0:
                    _disc_pct = (_promo_val / _sub_val) * 100
                    if _disc_pct > 15:
                        st.warning(f"⚠️ 折扣率 {_disc_pct:.1f}% 超过 15%，请确认是否正确")
                    else:
                        st.success(f"✅ 折扣率 {_disc_pct:.1f}%（≤ 15%，正常范围）")
            except:
                pass
            d["Total (RM)"]    = st.text_input("Total (RM)",    str(d.get("Total (RM)","")))

        # ══════════════════════════════════════════════════════════════
        # 🎯 SECTION: Auto-Suggest Product Items  (v12)
        # ══════════════════════════════════════════════════════════════
        st.markdown("---")
        label_color = "#6B5B95" if is_ast else "#1A1A1A"
        co_label    = "AI 服务套餐" if is_ast else "IGZ 服务套餐"
        st.markdown(
            f"<div class='section-title' style='color:{label_color}'>🎯 产品组合 — 自动匹配收据金额</div>",
            unsafe_allow_html=True)

        # Receipt total (locked from OCR)
        receipt_total = 0.0
        try:
            receipt_total = float(str(d.get("Total (RM)", 0)).replace(",", ""))
        except:
            pass

        # ── MODE selector ─────────────────────────────────────────────
        mode_opts = ["🤖 自动选品（Auto-Select）", "📦 手动选套餐（Manual Package）"]
        cur_mode  = st.session_state.get("item_select_mode", mode_opts[0])
        if cur_mode not in mode_opts:
            cur_mode = mode_opts[0]
        item_mode = st.radio("选择模式", mode_opts,
                             index=mode_opts.index(cur_mode),
                             horizontal=True, label_visibility="collapsed")
        st.session_state["item_select_mode"] = item_mode

        # ══════════════════════════════════════════════════════════════
        # MODE A: AUTO-SELECT — algorithm picks items that sum > card_total
        #         then sets Promo Rebate to make final = card_total exactly
        # ══════════════════════════════════════════════════════════════
        if item_mode == mode_opts[0]:
            if receipt_total > 0:
                # Run or re-use cached suggestion
                cache_key = f"auto_suggest_{receipt_total}_{company}"
                if st.session_state.get("auto_suggest_cache_key") != cache_key:
                    suggested_items, suggested_promo = auto_select_items(
                        receipt_total, company)
                    st.session_state["auto_suggest_items"]     = suggested_items
                    st.session_state["auto_suggest_promo"]     = suggested_promo
                    st.session_state["auto_suggest_cache_key"] = cache_key
                else:
                    suggested_items = st.session_state.get("auto_suggest_items", [])
                    suggested_promo = st.session_state.get("auto_suggest_promo", 0.0)

                if suggested_items:
                    subtotal_auto = sum(p for _, p in suggested_items)
                    promo_auto    = suggested_promo   # negative value
                    final_auto    = subtotal_auto + promo_auto  # = card_total

                    # ── Display suggestion box ─────────────────────────────
                    st.markdown(f"""
                    <div style='background:#0d1a0d;border-radius:12px;padding:16px 20px;
                                border-left:4px solid #22c55e;margin:10px 0'>
                        <div style='color:#86efac;font-size:11px;margin-bottom:6px'>
                            🤖 AUTO-SELECTED — 系统自动选品，Subtotal 稍高于刷卡金额，
                            通过 Promo Rebate 精确还原
                        </div>
                        <div style='color:#fff;font-size:20px;font-weight:900'>
                            RM {subtotal_auto:,.2f}
                            <span style='color:#f87171;font-size:14px;margin-left:8px'>
                                Promo {promo_auto:+,.2f}
                            </span>
                            <span style='color:#4ade80;font-size:14px;margin-left:8px'>
                                = RM {final_auto:,.2f} ✅
                            </span>
                        </div>
                    </div>""", unsafe_allow_html=True)

                    # ── Item breakdown ─────────────────────────────────────
                    st.markdown("**选中产品明细：**")
                    for name, price in suggested_items:
                        clean = re.sub(r"^\[[A-Z]-\d{3}\]\s*", "", name)
                        st.markdown(
                            f"<div style='padding:5px 12px;margin:2px 0;border-radius:6px;"
                            f"background:#0a1a0a;color:#86efac;font-size:13px'>"
                            f"✅ {clean} &nbsp;&nbsp;"
                            f"<span style='color:#4ade80;font-weight:700'>RM {price:,.2f}</span>"
                            f"</div>",
                            unsafe_allow_html=True)

                    # ── Write values to d ──────────────────────────────────
                    # Join item names as comma-separated for Product Item column
                    item_names_joined = "; ".join(
                        re.sub(r"^\[[A-Z]-\d{3}\]\s*", "", n)
                        for n, p in suggested_items)
                    d["Product Item"]       = item_names_joined
                    d["Qty"]                = len(suggested_items)
                    d["Unit Price (RM)"]    = f"{subtotal_auto / len(suggested_items):.2f}"
                    d["Total Amount (RM)"]  = f"{subtotal_auto:.2f}"
                    d["Subtotal (RM)"]      = f"{subtotal_auto:.2f}"
                    d["Promo Rebate (RM)"]  = f"{promo_auto:.2f}"   # e.g. "-112.00"
                    d["Total (RM)"]         = f"{final_auto:.2f}"

                    # Store for PDF (each item as separate line)
                    st.session_state["pkg_sub_items"] = [n for n, p in suggested_items]
                    st.session_state["auto_items_for_pdf"] = [
                        {"desc": re.sub(r"^\[[A-Z]-\d{3}\]\s*", "", n),
                         "unit_price": float(p), "qty": 1, "amount": float(p)}
                        for n, p in suggested_items
                    ]

                    # ── Refresh button ──────────────────────────────────────
                    if st.button("🔄 重新选品（换一组组合）", key="btn_refresh_auto"):
                        st.session_state.pop("auto_suggest_cache_key", None)
                        st.rerun()

                else:
                    st.warning("⚠️ 当前产品目录中无法找到合适的组合，请切换到手动选套餐模式。")
                    st.session_state["pkg_sub_items"] = []
                    st.session_state.pop("auto_items_for_pdf", None)
                    d["Product Item"] = ""
            else:
                st.info("ℹ️ 请先确认 Total (RM) 金额，系统将据此自动选品。")
                st.session_state["pkg_sub_items"] = []
                st.session_state.pop("auto_items_for_pdf", None)

        # ══════════════════════════════════════════════════════════════
        # MODE B: MANUAL PACKAGE — original dropdown selector
        # ══════════════════════════════════════════════════════════════
        else:
            packages  = _packages(company)
            pkg_names = list(packages.keys())
            cur_pkg   = st.session_state.get("selected_package", "")
            idx_pkg   = pkg_names.index(cur_pkg) if cur_pkg in pkg_names else 0

            selected_pkg = st.selectbox(
                f"选择套餐（{co_label}）",
                pkg_names,
                index=idx_pkg,
                help="选择套餐后自动带出包含项目，总价 = 收据金额")
            st.session_state["selected_package"] = selected_pkg

            sub_items = packages.get(selected_pkg, []) if selected_pkg else []

            if selected_pkg:
                st.markdown(f"""
                <div style='background:#1e1e2e;border-radius:10px;padding:14px 18px;
                            border-left:4px solid {label_color};margin:10px 0'>
                    <div style='color:#aaa;font-size:11px;margin-bottom:4px'>📦 PACKAGE</div>
                    <div style='color:#fff;font-size:15px;font-weight:700'>{selected_pkg}</div>
                    <div style='color:{label_color};font-size:18px;font-weight:800;margin-top:6px'>
                        RM {receipt_total:,.2f}
                        <span style='color:#666;font-size:12px;margin-left:8px'>= 收据金额 ✅</span>
                    </div>
                </div>""", unsafe_allow_html=True)

                if sub_items:
                    st.markdown("**包含项目 (Including):**")
                    for item in sub_items:
                        clean = re.sub(r"^\[[A-Z]-\d{3}\]\s*", "", item)
                        st.markdown(
                            f"<div style='padding:5px 12px;margin:2px 0;border-radius:6px;"
                            f"background:#0d0d1a;color:#888;font-size:13px'>"
                            f"↳ {clean} &nbsp;&nbsp;"
                            f"<span style='color:#444'>RM 0.00</span></div>",
                            unsafe_allow_html=True)
                else:
                    st.info("ℹ️ Custom Package — 子项目将由您手动填写在 Remarks 栏")

                d["Product Item"]      = selected_pkg
                d["Qty"]               = 1
                d["Unit Price (RM)"]   = f"{receipt_total:.2f}"
                d["Total Amount (RM)"] = f"{receipt_total:.2f}"
                st.session_state["pkg_sub_items"]    = sub_items
                st.session_state.pop("auto_items_for_pdf", None)
            else:
                st.session_state["pkg_sub_items"] = []
                st.session_state.pop("auto_items_for_pdf", None)
                d["Product Item"] = ""
                d["Unit Price (RM)"] = ""
                d["Total Amount (RM)"] = ""

        d["Remarks"] = st.text_area("Remarks", d.get("Remarks", ""), height=60)
        st.session_state.parsed_data = d

        # ── STEP 3: Generate Invoice ──────────────────────────────
        st.markdown("---")
        st.markdown("<div class='section-title'>🚀 Step 3 · 生成 Invoice</div>",
                    unsafe_allow_html=True)

        if not st.session_state.invoice_done:
            if st.button("📄  GENERATE INVOICE  —  生成收据",
                         type="primary", use_container_width=True):
                with st.spinner("⚙️ 正在生成 Invoice PDF…"):
                    try:
                        inv_no = d.get("Invoice No") or datetime.datetime.now().strftime("%Y%m%d%H%M")
                        receipt_total = float(str(d.get("Total (RM)","0")).replace(",",""))
                        pkg_name  = d.get("Product Item", "Services Rendered") or "Services Rendered"
                        sub_items = st.session_state.get("pkg_sub_items", [])

                        # ── Use auto-selected items if available (Mode A) ────
                        auto_pdf_items = st.session_state.get("auto_items_for_pdf", [])
                        if auto_pdf_items:
                            # Each item is already a dict with desc/unit_price/qty/amount
                            items_list = auto_pdf_items
                        else:
                            # Mode B: single package line + sub-items at RM 0.00
                            items_list = [{
                                "desc":       pkg_name,
                                "unit_price": receipt_total,
                                "qty":        1,
                                "amount":     receipt_total,
                                "is_package": True,
                            }]
                            for si in sub_items:
                                clean_si = re.sub(r"^\[[A-Z]-\d{3}\]\s*", "", si)
                                items_list.append({
                                    "desc":       f"↳ {clean_si}",
                                    "unit_price": 0.0,
                                    "qty":        1,
                                    "amount":     0.0,
                                    "is_subitem": True,
                                })

                        pdf_p    = generate_invoice_pdf(company, d, inv_no, items_list)
                        inv_link = str(pdf_p)

                        COLS = get_cols(company)
                        for col in COLS:
                            if col not in ("Receipt Link","Invoice Link"):
                                update_cell(wb, ws, st.session_state.excel_row, col,
                                            d.get(col,""), excel_path, company)
                        update_cell(wb, ws, st.session_state.excel_row,
                                    "Invoice Link", inv_link, excel_path, company)

                        st.session_state.invoice_path = str(pdf_p)
                        st.session_state.invoice_done = True
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ 生成失败: {e}")
        else:
            st.success("✅ Invoice 已生成！")
            inv_path = st.session_state.invoice_path
            with open(inv_path,"rb") as f:
                st.download_button("⬇️ 下载 Invoice PDF", data=f,
                                   file_name=Path(inv_path).name,
                                   mime="application/pdf",
                                   use_container_width=True)
            # PNG preview
            png_p = str(Path(inv_path).with_suffix("")) + "_prev.png"
            try:
                subprocess.run(["pdftoppm","-r","150","-png","-singlefile",
                                inv_path, png_p.replace(".png","")],
                               check=True, capture_output=True)
                if os.path.exists(png_p):
                    st.image(png_p, caption="Invoice Preview", use_container_width=True)
            except: pass

            if st.button("📝 新增下一张收据"):
                for k in ["parsed_data","receipt_path","excel_row",
                          "invoice_done","invoice_path","items_list"]:
                    st.session_state[k] = {} if k=="parsed_data" else ([] if k=="items_list" else None)
                st.session_state.invoice_done = False
                # Reset package selection for new receipt
                st.session_state["selected_package"] = ""
                st.session_state["pkg_sub_items"] = []
                st.rerun()

    # ── STEP 4: Excel summary table ───────────────────────────────
    st.markdown("---")
    st.markdown(f"<div class='section-title'>📊 汇总表格 — {company}</div>",
                unsafe_allow_html=True)

    with open(excel_path,"rb") as f:
        st.download_button(
            label=f"⬇️ 下载 Excel 汇总表 ({Path(excel_path).name})",
            data=f, file_name=Path(excel_path).name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    import pandas as pd
    COLS = get_cols(company)
    lc   = "6B5B95" if is_ast else "1A1A1A"
    COL_W = {
        "Invoice No":130,"Date":110,"Due Date":110,"Bill To":120,"Company":160,
        "Type":140,"Card No":175,"Approval Code":115,
        "Receipt No":105,"Ref No":150,
        "Product Item":310,
        "Qty":65,"Unit Price (RM)":120,"Total Amount (RM)":135,
        "Subtotal (RM)":120,"Promo Rebate (RM)":140,"Tax":65,
        "Total (RM)":110,"Remarks":200,
        "Receipt Link":280,"Invoice Link":280,
    }

    try:
        df = pd.read_excel(excel_path, skiprows=1, header=0)
        col_cfg = {}
        for col in df.columns:
            w = COL_W.get(col, 120)
            if col in ("Receipt Link","Invoice Link"):
                col_cfg[col] = st.column_config.LinkColumn(col, width=w, display_text="🔗 查看")
            elif col in ("Subtotal (RM)","Unit Price (RM)","Total Amount (RM)","Total (RM)",
                         "Promo Rebate (RM)"):
                col_cfg[col] = st.column_config.NumberColumn(col, width=w, format="RM %.2f")
            elif col == "Qty":
                col_cfg[col] = st.column_config.NumberColumn(col, width=w, format="%d")
            else:
                col_cfg[col] = st.column_config.TextColumn(col, width=w)

        st.markdown(
            f"<p style='font-size:11px;color:#888'>共 {len(df)} 条记录 · "
            f"{len(df.columns)} 个字段 · 👉 <b>横向滑动查看全部列</b></p>",
            unsafe_allow_html=True)

        st.dataframe(df, column_config=col_cfg,
                     use_container_width=False,
                     width=2200,
                     height=420,
                     hide_index=True)
    except Exception as e:
        st.warning(f"暂无记录: {e}")

    st.caption(f"📁 {excel_path}  ·  Invoices: {INVOICES_DIR}")


if __name__ == "__main__":
    main()
